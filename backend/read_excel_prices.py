#!/usr/bin/env python3
"""
Script to read the FINAL Supplies Table Excel file and generate SQL updates
This helps update inventory prices from the Excel file
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages")
    import openpyxl

def read_excel_and_generate_updates():
    excel_path = '/Users/francescoassalone/Desktop/FINAL Supplies Table.xlsx'
    
    print("="*80)
    print("READING EXCEL FILE: FINAL Supplies Table.xlsx")
    print("="*80)
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        print(f"Sheet name: {sheet.title}")
        print(f"Total rows: {sheet.max_row}")
        print(f"Total columns: {sheet.max_column}\n")
        
        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        print(f"Columns found: {headers}\n")
        print("="*80)
        print("EXCEL DATA PREVIEW (First 10 rows)")
        print("="*80)
        
        # Show first 10 rows
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=11, values_only=True), 1):
            print(f"Row {i}: {row}")
        
        print("\n" + "="*80)
        print("GENERATING SQL UPDATE STATEMENTS")
        print("="*80)
        
        # Generate SQL updates
        sql_updates = []
        items_found = []
        
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            # Try to extract item name and price from various column positions
            item_name = None
            price = None
            
            # Common patterns: first column is usually name, look for price in following columns
            if row[0]:  # First column should be item name
                item_name = str(row[0]).strip()
            
            # Look for price (a number) in subsequent columns
            for cell_value in row[1:]:
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                    price = float(cell_value)
                    break
            
            if item_name and price and price > 0:
                items_found.append((item_name, price))
                sql_updates.append(f"UPDATE inventory_items SET unit_price = {price} WHERE item_name = '{item_name.replace(\"'\", \"''\")}';")
        
        print(f"\nFound {len(items_found)} items with prices:\n")
        
        for name, price in items_found[:20]:  # Show first 20
            print(f"  {name}: ${price:.2f}")
        
        if len(items_found) > 20:
            print(f"  ... and {len(items_found) - 20} more items\n")
        
        # Save SQL to file
        sql_file_path = '/Users/francescoassalone/Desktop/inventory-management-system/backend/update_prices.sql'
        with open(sql_file_path, 'w') as f:
            f.write("-- SQL UPDATE STATEMENTS FOR INVENTORY PRICES\n")
            f.write("-- Generated from FINAL Supplies Table.xlsx\n\n")
            for sql in sql_updates:
                f.write(sql + "\n")
        
        print("="*80)
        print(f"✓ SQL file saved to: {sql_file_path}")
        print("="*80)
        print("\nTo apply these updates:")
        print("1. Review the SQL file to make sure it looks correct")
        print("2. Run: node src/updatePricesFromExcel.js")
        print("   OR manually execute the SQL statements")
        
        return items_found
        
    except FileNotFoundError:
        print(f"ERROR: Excel file not found at: {excel_path}")
        print("Please make sure the file exists at this location.")
        return []
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

if __name__ == "__main__":
    read_excel_and_generate_updates()
